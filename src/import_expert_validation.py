#!/usr/bin/env python3
"""Convert the expert XLSX workbook into canonical review CSV.

The expert workbook is intentionally not modified. This adapter maps the human-
friendly Dutch columns to the validation pipeline's canonical field names.
Proposed corrections are never auto-applied.
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path
from openpyxl import load_workbook

HEADER_MAP = {
    "Object-ID": "object_id",
    "Beoordeling*": "review_decision",
    "Voorgestelde correctie": "proposed_correction",
    "Toelichting / reden": "review_comment",
    "Naam reviewer": "reviewer",
    "Datum review": "review_date",
    "Bronpassage / gegenereerde inhoud": "reviewed_text",
    "Operator": "reviewed_operator",
    "Drempel": "reviewed_threshold",
    "Eenheid": "reviewed_unit",
    "Scorepunten": "reviewed_score_points",
}
ALLOWED = {"", "approve", "revise", "reject"}


def norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_xlsx(path: Path, sheet: str = "Validatie") -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    missing = [h for h in HEADER_MAP if h not in idx]
    if missing:
        raise ValueError(f"Missing required workbook columns: {missing}")

    rows = []
    for cells in ws.iter_rows(min_row=2):
        values = [c.value for c in cells]
        oid = norm(values[idx["Object-ID"]])
        if not oid:
            continue
        row = {out: norm(values[idx[src]]) for src, out in HEADER_MAP.items()}
        row["review_decision"] = row["review_decision"].lower()
        if row["review_decision"] not in ALLOWED:
            raise ValueError(f"Invalid decision for {oid}: {row['review_decision']!r}")
        rows.append(row)
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--sheet", default="Validatie")
    args = ap.parse_args()

    rows = load_xlsx(args.xlsx, args.sheet)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    fields = ["object_id", "review_decision", "proposed_correction", "review_comment", "reviewer", "review_date", "reviewed_text", "reviewed_operator", "reviewed_threshold", "reviewed_unit", "reviewed_score_points"]
    with args.out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        w.writerows(rows)
    print(f"converted_rows={len(rows)} output={args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
